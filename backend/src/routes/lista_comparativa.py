from flask import Blueprint, request, jsonify
from src.db import mongo
from bson.objectid import ObjectId
from datetime import datetime
import openpyxl
import os
from werkzeug.utils import secure_filename

lista_comparativa_bp = Blueprint('lista_comparativa', __name__)

@lista_comparativa_bp.route('/lista-comparativa/buscar', methods=['GET'])
def buscar_en_proveedores():
    """Buscar medicamentos en todas las listas de proveedores con descuentos aplicados"""
    try:
        query = request.args.get('q', '').strip()
        proveedor_id = request.args.get('proveedor_id')
        
        if not query:
            return jsonify([])

        pipeline = []
        match_filter = {
            '$or': [
                {'codigo': {'$regex': query, '$options': 'i'}},
                {'descripcion': {'$regex': query, '$options': 'i'}},
                {'laboratorio': {'$regex': query, '$options': 'i'}}
            ],
            'disponible': True
        }
        if proveedor_id:
            match_filter['proveedor_id'] = ObjectId(proveedor_id)
        
        pipeline.append({'$match': match_filter})

        pipeline.extend([
            {
                '$lookup': {
                    'from': 'proveedores',
                    'localField': 'proveedor_id',
                    'foreignField': '_id',
                    'as': 'proveedor_info'
                }
            },
            {'$unwind': '$proveedor_info'},
            {
                '$addFields': {
                    'precio_original': {'$ifNull': ['$precio_descuento', '$precio']}
                }
            },
            {
                '$addFields': {
                    'descuento_aplicado': {'$multiply': ['$precio_original', {'$divide': ['$proveedor_info.descuento_comercial', 100]}]},
                }
            },
            {
                '$addFields': {
                    'precio_final_comercial': {'$subtract': ['$precio_original', '$descuento_aplicado']}
                }
            },
            {
                '$addFields': {
                    'descuento_pronto_pago_aplicado': {'$multiply': ['$precio_final_comercial', {'$divide': ['$proveedor_info.descuento_pronto_pago', 100]}]},
                }
            },
            {
                '$addFields': {
                    'precio_final_pronto_pago': {'$subtract': ['$precio_final_comercial', '$descuento_pronto_pago_aplicado']}
                }
            },
            {
                '$group': {
                    '_id': {'codigo': '$codigo', 'descripcion': '$descripcion'},
                    'laboratorio': {'$first': '$laboratorio'},
                    'proveedores': {
                        '$push': {
                            'proveedor_id': '$proveedor_id',
                            'proveedor_nombre': '$proveedor_info.nombre',
                            'precio_original': '$precio_original',
                            'descuento_comercial': '$proveedor_info.descuento_comercial',
                            'descuento_pronto_pago': '$proveedor_info.descuento_pronto_pago',
                            'precio_con_descuento_comercial': '$precio_final_comercial',
                            'precio_con_descuento_total': '$precio_final_pronto_pago',
                            'ahorro_comercial': '$descuento_aplicado',
                            'ahorro_pronto_pago': '$descuento_pronto_pago_aplicado',
                            'ahorro_total': {'$add': [
                                '$descuento_aplicado',
                                '$descuento_pronto_pago_aplicado'
                            ]},
                            'dias_credito': '$proveedor_info.dias_credito',
                            'fecha_actualizacion': '$fecha_actualizacion'
                        }
                    }
                }
            },
            {
                '$project': {
                    '_id': 0,
                    'codigo': '$_id.codigo',
                    'descripcion': '$_id.descripcion',
                    'laboratorio': 1,
                    'proveedores': 1
                }
            }
        ])

        productos_agrupados = list(mongo.db.lista_proveedor.aggregate(pipeline))

        for producto in productos_agrupados:
            producto['proveedores'].sort(key=lambda x: x['precio_con_descuento_comercial'])
            if producto['proveedores']:
                producto['proveedores'][0]['es_mejor_precio'] = True
                producto['mejor_precio'] = producto['proveedores'][0]['precio_con_descuento_comercial']
                producto['mejor_proveedor'] = producto['proveedores'][0]['proveedor_nombre']
                producto['ahorro_mejor_opcion'] = producto['proveedores'][0]['ahorro_total']
                for i, proveedor in enumerate(producto['proveedores'][1:], 1):
                    diferencia = proveedor['precio_con_descuento_comercial'] - producto['mejor_precio']
                    proveedor['diferencia_con_mejor'] = diferencia
                    proveedor['porcentaje_diferencia'] = (diferencia / producto['mejor_precio']) * 100 if producto['mejor_precio'] > 0 else 0

        resultado_final = sorted(productos_agrupados, key=lambda x: x.get('mejor_precio', float('inf')))

        return jsonify(resultado_final)
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@lista_comparativa_bp.route('/lista-comparativa/proveedores', methods=['GET'])
def get_proveedores_con_listas():
    """Obtener proveedores que tienen listas de precios"""
    try:
        pipeline = [
            {'$group': {'_id': '$proveedor_id'}},
            {'$lookup': {
                'from': 'proveedores',
                'localField': '_id',
                'foreignField': '_id',
                'as': 'proveedor_info'
            }},
            {'$unwind': '$proveedor_info'},
            {'$replaceRoot': {'newRoot': '$proveedor_info'}}
        ]
        proveedores = list(mongo.db.lista_proveedor.aggregate(pipeline))
        return jsonify([{
            'id': str(p['_id']),
            'nombre': p['nombre'],
            'contacto': p.get('contacto'),
            'total_productos': mongo.db.lista_proveedor.count_documents({'proveedor_id': p['_id']})
        } for p in proveedores])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@lista_comparativa_bp.route('/lista-comparativa/upload/<proveedor_id>', methods=['POST'])
def upload_lista_proveedor(proveedor_id):
    """Subir lista de precios de un proveedor desde Excel"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No se encontró archivo'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No se seleccionó archivo'}), 400
        
        proveedor = mongo.db.proveedores.find_one_or_404({'_id': ObjectId(proveedor_id)})
        
        filename = secure_filename(file.filename)
        temp_path = os.path.join('/tmp', filename)
        file.save(temp_path)
        
        wb = openpyxl.load_workbook(temp_path)
        ws = wb.active
        
        productos_procesados = 0
        productos_actualizados = 0
        errores = []
        
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=2, values_only=True):
            if any(cell for cell in row if cell and str(cell).strip()):
                headers = [str(cell).strip().upper() if cell else '' for cell in row]
                break
        
        col_mapping = {}
        for i, header in enumerate(headers):
            if 'CODIGO' in header or 'COD' in header:
                col_mapping['codigo'] = i
            elif 'DESCRIPCION' in header or 'PRODUCTO' in header or 'NOMBRE' in header:
                col_mapping['descripcion'] = i
            elif 'LABORATORIO' in header or 'LAB' in header:
                col_mapping['laboratorio'] = i
            elif 'PRECIO' in header and 'DESCUENTO' not in header:
                col_mapping['precio'] = i
            elif 'DESCUENTO' in header or 'PRECIO_DESCUENTO' in header:
                col_mapping['precio_descuento'] = i
        
        for row_num, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            try:
                if not any(cell for cell in row if cell):
                    continue
                
                codigo = str(row[col_mapping.get('codigo', 0)] or '').strip()
                descripcion = str(row[col_mapping.get('descripcion', 1)] or '').strip()
                laboratorio = str(row[col_mapping.get('laboratorio', 2)] or '').strip()
                
                if not codigo or not descripcion:
                    continue
                
                try:
                    precio = float(row[col_mapping.get('precio', 3)] or 0)
                    precio_descuento = None
                    if 'precio_descuento' in col_mapping and row[col_mapping['precio_descuento']]:
                        precio_descuento = float(row[col_mapping['precio_descuento']])
                except (ValueError, TypeError):
                    errores.append(f"Fila {row_num}: Error en formato de precio")
                    continue
                
                update_data = {
                    'descripcion': descripcion,
                    'laboratorio': laboratorio,
                    'precio': precio,
                    'precio_descuento': precio_descuento,
                    'fecha_actualizacion': datetime.utcnow(),
                    'proveedor_id': ObjectId(proveedor_id)
                }

                result = mongo.db.lista_proveedor.update_one(
                    {'proveedor_id': ObjectId(proveedor_id), 'codigo': codigo},
                    {'$set': update_data},
                    upsert=True
                )

                if result.upserted_id:
                    productos_procesados += 1
                else:
                    productos_actualizados += 1
                
            except Exception as e:
                errores.append(f"Fila {row_num}: {str(e)}")
                continue
        
        os.remove(temp_path)
        
        return jsonify({
            'message': 'Lista de precios procesada exitosamente',
            'productos_nuevos': productos_procesados,
            'productos_actualizados': productos_actualizados,
            'errores': errores[:10]
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

